from openpyxl import Workbook
from openpyxl import load_workbook
from datetime import datetime

timestamp = datetime.now().strftime("%d.%m.%Y-%H.%M")

wb = Workbook()

ws = wb.active
ws.title = "tesuto"

target = wb.copy_worksheet(ws)

print(wb.sheetnames)

a = ws['A1']
a.value = "Heading"

fill = ws.cell(row=2, column=1, value = 6000)

#wb.save(timestamp + '.xlsx')

wb1 = load_workbook("HDI vs GDP per capita.xlsx")
ws1 = wb1.active
first = ws1["A1":"A5"]
print([i[0].value for i in first])