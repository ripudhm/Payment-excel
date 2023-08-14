
import pickle
import os.path
from apiclient import errors
import email
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from googleapiclient.discovery import build
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
import base64
from datetime import datetime
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side

# If modifying these scopes, delete the file token.pickle.
SCOPES = ['https://mail.google.com/']
timestamp = datetime.now().strftime("%d.%m.%Y-%H.%M")
#rowc = 1
#colc = 1
wb = Workbook()
ws = wb.active
ws.title = "test"

def get_service():
    """
    Authenticate the google api client and return the service object 
    to make further calls

    PARAMS
        None

    RETURNS
        service api object from gmail for making calls
    """
    creds = None

    # The file token.pickle stores the user's access and refresh tokens, and is
    # created automatically when the authorization flow completes for the first
    # time.
    if os.path.exists('token.pickle'):
        with open('token.pickle', 'rb') as token:
            creds = pickle.load(token)

    # If there are no (valid) credentials available, let the user log in.
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file('credentials.json', SCOPES)
            creds = flow.run_local_server(port=0)

        # Save the credentials for the next run
        with open('token.pickle', 'wb') as token:
            pickle.dump(creds, token)


    service = build('gmail', 'v1', credentials=creds)

    return service


def search_message(service, user_id, search_string):
    """
    Search the inbox for emails using standard gmail search parameters
    and return a list of email IDs for each result

    PARAMS:
        service: the google api service object already instantiated
        user_id: user id for google api service ('me' works here if
        already authenticated)
        search_string: search operators you can use with Gmail
        (see https://support.google.com/mail/answer/7190?hl=en for a list)

    RETURNS:
        List containing email IDs of search query
    """
    try:
        # initiate the list for returning
        list_ids = []

        # get the id of all messages that are in the search string
        search_ids = service.users().messages().list(userId=user_id, q=search_string).execute()
        
        # if there were no results, print warning and return empty string
        try:
            ids = search_ids['messages']

        except KeyError:
            print("WARNING: the search queried returned 0 results")
            print("returning an empty string")
            return ""

        if len(ids)>1:
            for msg_id in ids:
                list_ids.append(msg_id['id'])
            return(list_ids)

        else:
            list_ids.append(ids['id'])
            return list_ids
        
    except (errors.HttpError, error):
        print("An error occured: %s") % error


def get_message(service, user_id, msg_id):
    """
    Search the inbox for specific message by ID and return it back as a 
    clean string. String may contain Python escape characters for newline
    and return line. 
    
    PARAMS
        service: the google api service object already instantiated
        user_id: user id for google api service ('me' works here if
        already authenticated)
        msg_id: the unique id of the email you need

    RETURNS
        A string of encoded text containing the message body
    """
    try:
        # grab the message instance
        message = service.users().messages().get(userId=user_id, id=msg_id,format='raw').execute()

        # decode the raw string, ASCII works pretty well here
        msg_str = base64.urlsafe_b64decode(message['raw'].encode('ASCII'))

        # grab the string from the byte object
        mime_msg = email.message_from_bytes(msg_str)

        # check if the content is multipart (it usually is)
        content_type = mime_msg.get_content_maintype()
        if content_type == 'multipart':
            # there will usually be 2 parts the first will be the body in text
            # the second will be the text in html
            parts = mime_msg.get_payload()

            # return the encoded text
            final_content = parts[0].get_payload()
            return final_content

        elif content_type == 'text':
            return mime_msg.get_payload()

        else:
            return ""
            print("\nMessage is not text or multipart, returned an empty string")
    # unsure why the usual exception doesn't work in this case, but 
    # having a standard Exception seems to do the trick
    except Exception:
        print("An error occured: %s") % error

def get_subject(service, user_id, msg_id):
    try:
        # grab the message instance
        message = service.users().messages().get(userId=user_id, id=msg_id).execute()
        payld = message['payload']
        headr = payld['headers']
        sub = [i['value'] for i in headr if i['name'] == "Subject"]
        return sub[0]
    except Exception:
        print("An error occured: %s") % error

def get_all_unread_in_primary(service, user_id):
    try:
        unread_in_primary_list = []
        ids = service.users().messages().list(userId = user_id, labelIds = ["UNREAD", "CATEGORY_PERSONAL"]).execute()
        id_list = ids['messages']
        for i in id_list:
            unread_in_primary_list.append(i['id'])
        #print(unread_in_primary_list)
        fin_dct = {}
        service = get_service()
        
        for i in unread_in_primary_list:
            fin_dct[i] = get_subject(service, 'me', i)

        rowc = 2
        colc = 1
        ws["A1"].value = "Message ID"
        ws["A1"].font = Font(bold = True)
        ws["A1"].border = Border(bottom=Side(border_style="thick", \
            color="000000"))
        ws["B1"].value = "Subject"
        ws["B1"].font = Font(bold = True)
        ws["B1"].border = Border(bottom=Side(border_style="thick", \
            color="000000"))

        for msg_id, sub in fin_dct.items():
            fill1 = ws.cell(row=rowc, column=colc, value = msg_id)
            fill2 = ws.cell(row=rowc, column=colc + 1, value = sub)
            rowc += 1

        wb.save(timestamp + ".xlsx")
        return fin_dct
    except Exception:
        print("An error occured: %s") % error

def get_filtered_body_and_subject(service, user_id, filter_str):
    try:
        final_list_of_bodysub = []
        with open(timestamp + ".txt", 'w') as f:
            for i, sub in get_all_unread_in_primary(service, 'me').items():
                temp_dct = {}
                if filter_str in sub:
                    temp_dct['msg_id'] = i
                    temp_dct['Subject'] = sub
                    temp_dct['Body'] = get_message(service, 'me', i)
                    final_list_of_bodysub.append(temp_dct)
                f.write(i + "\n" + sub + "\n" + get_message(service, 'me', i) + "\n" + "\n")
        return final_list_of_bodysub
    except Exception:
        print("An error occured: %s") % error

def mark_as_read(service, msg_id):
    service.users().messages().modify(userId = 'me', id = str(msg_id), body= {'addLabelIds': [], \
        'removeLabelIds': ["UNREAD"]}).execute()

def forward(service, msg_id, to):
    message = get_message(service, 'me', msg_id)
    subject = get_subject(service, 'me', msg_id)
    #return message

    emailMsg = message
    mimeMessage = MIMEMultipart()
    mimeMessage['to'] = to
    mimeMessage['subject'] = subject
    mimeMessage.attach(MIMEText(emailMsg, 'plain'))
    raw_string = base64.urlsafe_b64encode(mimeMessage.as_bytes()).decode()

    mark_as_read(service, msg_id)

    sent_mes = service.users().messages().send(userId = 'me', body = {'raw':raw_string}).execute()
    print(sent_mes)








